import os
import logging
from datetime import datetime
from contextlib import contextmanager
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

import pandas as pd
import psycopg2
from psycopg2.pool import ThreadedConnectionPool
from psycopg2.extras import RealDictCursor
from flask import (
    Flask, render_template, request, redirect, url_for, flash, jsonify,
    send_file
)
from io import BytesIO
from openpyxl.utils import get_column_letter

app = Flask(__name__)
# Lee SECRET_KEY desde entorno; deja fallback para desarrollo local
app.secret_key = os.getenv('SECRET_KEY', 'dev-only-change-in-prod')
DATA_DIR = 'data'
os.makedirs(DATA_DIR, exist_ok=True)

# =========================
#  Conexión a Postgres/Neon
# =========================

def normalize_db_url(raw_url: str) -> str:
    """
    - Asegura sslmode=require.
    - Elimina channel_binding (puede romper según cliente).
    - Agrega application_name.
    """
    if not raw_url:
        return raw_url
    parsed = urlparse(raw_url)
    q = dict(parse_qsl(parsed.query, keep_blank_values=True))
    q.pop("channel_binding", None)
    if "sslmode" not in q:
        q["sslmode"] = "require"
    if "application_name" not in q:
        q["application_name"] = "mensajeria_plataf"
    new_query = urlencode(q)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))

DATABASE_URL = normalize_db_url(os.getenv("DATABASE_URL", ""))

if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL no está definida. En Render pon tu URL POOLER de Neon en Environment Variables.")

POOL_MIN = int(os.getenv("PG_POOL_MIN", "1"))
POOL_MAX = int(os.getenv("PG_POOL_MAX", "5"))

pool = ThreadedConnectionPool(minconn=POOL_MIN, maxconn=POOL_MAX, dsn=DATABASE_URL)

@contextmanager
def get_conn():
    conn = pool.getconn()
    try:
        yield conn
        conn.commit()
    finally:
        pool.putconn(conn)

def db_exec(sql, params=()):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)

def db_fetchone_dict(sql, params=()):
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            return cur.fetchone()

def db_fetchall_dict(sql, params=()):
    with get_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            return cur.fetchall()

def read_sql_df(sql, params=None):
    # Pandas con psycopg2 puede mostrar un warning, pero funciona correctamente.
    with get_conn() as conn:
        return pd.read_sql(sql, conn, params=params)

# =========================
#   Esquema (si no existe)
# =========================

def ensure_schema():
    # Zonas / Mensajeros / Guías
    db_exec("""
        CREATE TABLE IF NOT EXISTS zonas (
            nombre TEXT PRIMARY KEY,
            tarifa NUMERIC NOT NULL
        );
    """)
    db_exec("""
        CREATE TABLE IF NOT EXISTS mensajeros (
            nombre TEXT PRIMARY KEY,
            zona   TEXT REFERENCES zonas(nombre)
        );
    """)
    db_exec("""
        CREATE TABLE IF NOT EXISTS guias (
            remitente   TEXT,
            numero_guia TEXT PRIMARY KEY,
            destinatario TEXT,
            direccion   TEXT,
            ciudad      TEXT
        );
    """)
    # Despachos / Recepciones
    db_exec("""
        CREATE TABLE IF NOT EXISTS despachos (
            numero_guia TEXT PRIMARY KEY,
            mensajero   TEXT,
            zona        TEXT,
            fecha       TIMESTAMPTZ NOT NULL
        );
    """)
    db_exec("""
        CREATE TABLE IF NOT EXISTS recepciones (
            numero_guia TEXT PRIMARY KEY,
            tipo        TEXT,           -- ENTREGADA / DEVUELTA
            motivo      TEXT,
            fecha       TIMESTAMPTZ NOT NULL
        );
    """)
    # Recogidas
    db_exec("""
        CREATE TABLE IF NOT EXISTS recogidas (
            id          SERIAL PRIMARY KEY,
            numero_guia TEXT,
            fecha       TIMESTAMPTZ NOT NULL,
            observaciones TEXT
        );
    """)

    # ==== NUEVO: Clientes y vínculo con Recogidas ====
    db_exec("""
        CREATE TABLE IF NOT EXISTS clientes (
            id SERIAL PRIMARY KEY,
            nombre TEXT UNIQUE NOT NULL,
            telefono TEXT,
            direccion TEXT,
            ciudad TEXT,
            contacto TEXT
        );
    """)
    # Columna cliente_id en recogidas (nullable)
    db_exec("""
        ALTER TABLE recogidas
        ADD COLUMN IF NOT EXISTS cliente_id INTEGER REFERENCES clientes(id);
    """)

    # Índices útiles
    db_exec("CREATE INDEX IF NOT EXISTS idx_mensajeros_zona ON mensajeros(zona);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_guias_numero ON guias(numero_guia);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_despachos_fecha ON despachos(fecha);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_recepciones_fecha ON recepciones(fecha);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_recogidas_fecha ON recogidas(fecha);")
    db_exec("CREATE INDEX IF NOT EXISTS idx_recogidas_cliente ON recogidas(cliente_id);")

# =========================
#   Modelos en memoria
# =========================

class Zona:
    def __init__(self, nombre, tarifa):
        self.nombre = nombre
        self.tarifa = float(tarifa) if tarifa is not None else 0.0

class Mensajero:
    def __init__(self, nombre, zona):
        self.nombre = nombre
        self.zona = zona

zonas = []
mensajeros = []
guias = pd.DataFrame(columns=['remitente', 'numero_guia', 'destinatario', 'direccion', 'ciudad'])
despachos = []
recepciones = []
recogidas = []
clientes = []  # cache liviano para selects

def cargar_datos_desde_db():
    global zonas, mensajeros, guias, despachos, recepciones, recogidas, clientes

    zrows = db_fetchall_dict("SELECT nombre, tarifa FROM zonas;")
    zonas = [Zona(r["nombre"], r["tarifa"]) for r in zrows]

    mrows = db_fetchall_dict("SELECT nombre, zona FROM mensajeros;")
    zonas_map = {z.nombre: z for z in zonas}
    mensajeros_local = []
    for r in mrows:
        zona_obj = zonas_map.get(r["zona"])
        mensajeros_local.append(Mensajero(r["nombre"], zona_obj))
    globals()["mensajeros"] = mensajeros_local

    guias_df = read_sql_df("SELECT remitente, numero_guia, destinatario, direccion, ciudad FROM guias;")
    globals()["guias"] = guias_df

    despachos_list = db_fetchall_dict("SELECT numero_guia, mensajero, zona, fecha FROM despachos ORDER BY fecha DESC;")
    recepciones_list = db_fetchall_dict("SELECT numero_guia, tipo, motivo, fecha FROM recepciones ORDER BY fecha DESC;")
    # incluye cliente_id
    recogidas_list = db_fetchall_dict("SELECT id, numero_guia, fecha, observaciones, cliente_id FROM recogidas ORDER BY fecha DESC;")
    clientes_list = db_fetchall_dict("SELECT id, nombre, telefono, direccion, ciudad, contacto FROM clientes ORDER BY nombre;")

    globals()["despachos"]  = despachos_list
    globals()["recepciones"] = recepciones_list
    globals()["recogidas"]   = recogidas_list
    globals()["clientes"]    = clientes_list

# Inicializa
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO").upper())
ensure_schema()
cargar_datos_desde_db()

# =========================
#   Util: Excel en memoria
# =========================

def df_to_excel_download(df: pd.DataFrame, base_name: str, sheet_name: str = "Hoja1"):
    """
    Retorna un send_file con un Excel generado en memoria.
    - Auto-anchos de columnas
    - Formato fecha si aplica
    """
    if df is None:
        df = pd.DataFrame()

    # Normaliza fechas a naive (excel-friendly)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)
            except Exception:
                pass

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet = sheet_name
        (df if not df.empty else pd.DataFrame(columns=list(df.columns))).to_excel(
            writer, index=False, sheet_name=sheet
        )

        ws = writer.sheets[sheet]
        # auto width
        for idx, col in enumerate(df.columns if len(df.columns) else [" "], start=1):
            if df.empty:
                max_len = len(str(col))
            else:
                max_len = max([len(str(col))] + [len(str(x)) for x in df[col].astype(str).values])
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(40, max_len + 2))

        # formato de fecha por nombre de columna típica
        for name in df.columns:
            if "fecha" in name.lower():
                col_idx = list(df.columns).index(name) + 1
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                    for cell in row:
                        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_{stamp}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
#          Rutas
# =========================

@app.route("/")
def index():
    return render_template('index.html')

@app.route("/cargar_base", methods=["GET", "POST"])
def cargar_base():
    global guias
    if request.method == 'POST':
        archivo = request.files.get('archivo_excel')
        if archivo:
            df = pd.read_excel(archivo)
            required_cols = ['remitente', 'numero_guia', 'destinatario', 'direccion', 'ciudad']
            if all(col in df.columns for col in required_cols):
                with get_conn() as conn:
                    cur = conn.cursor()
                    for _, row in df.iterrows():
                        numero = str(row['numero_guia'])
                        cur.execute("SELECT 1 FROM guias WHERE numero_guia = %s;", (numero,))
                        existe = cur.fetchone()
                        if not existe:
                            cur.execute("""
                                INSERT INTO guias(remitente, numero_guia, destinatario, direccion, ciudad)
                                VALUES (%s, %s, %s, %s, %s);
                            """, (row['remitente'], numero, row['destinatario'], row['direccion'], row['ciudad']))
                guias = read_sql_df("SELECT remitente, numero_guia, destinatario, direccion, ciudad FROM guias;")
                globals()["guias"] = guias

                # Guardar archivo (almacenamiento efímero en Render, válido para debug)
                archivo.save(os.path.join(DATA_DIR, archivo.filename))
                flash('Base de datos cargada correctamente.', 'success')
            else:
                flash('El archivo debe contener las columnas: ' + ", ".join(required_cols), 'danger')
    return render_template('cargar_base.html')

@app.route("/registrar_zona", methods=["GET", "POST"])
def registrar_zona():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        tarifa = request.form.get('tarifa')
        if nombre and tarifa:
            try:
                tarifa_float = float(tarifa)
                existe = db_fetchone_dict("SELECT 1 AS x FROM zonas WHERE nombre = %s;", (nombre,))
                if existe:
                    flash('La zona ya existe', 'warning')
                else:
                    db_exec("INSERT INTO zonas(nombre, tarifa) VALUES (%s, %s);", (nombre, tarifa_float))
                    flash(f'Zona {nombre} registrada con tarifa {tarifa_float}', 'success')
                cargar_datos_desde_db()
            except ValueError:
                flash('Tarifa inválida, debe ser un número', 'danger')
        else:
            flash('Debe completar todos los campos', 'danger')
    return render_template('registrar_zona.html', zonas=zonas)

@app.route("/registrar_mensajero", methods=["GET", "POST"])
def registrar_mensajero():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        zona_nombre = request.form.get('zona')
        if nombre and zona_nombre:
            zona_obj = next((z for z in zonas if z.nombre == zona_nombre), None)
            if not zona_obj:
                flash('Zona no encontrada', 'danger')
                return redirect(url_for('registrar_mensajero'))

            existe = db_fetchone_dict("SELECT 1 AS x FROM mensajeros WHERE nombre = %s;", (nombre,))
            if existe:
                flash('El mensajero ya existe', 'warning')
            else:
                db_exec("INSERT INTO mensajeros(nombre, zona) VALUES (%s, %s);", (nombre, zona_nombre))
                flash(f'Mensajero {nombre} registrado en zona {zona_nombre}', 'success')
            cargar_datos_desde_db()
        else:
            flash('Debe completar todos los campos', 'danger')
    return render_template('registrar_mensajero.html', zonas=zonas, mensajeros=mensajeros)

@app.route("/despachar_guias", methods=["GET", "POST"])
def despachar_guias():
    if request.method == 'POST':
        mensajero_nombre = request.form.get('mensajero')
        guias_input = request.form.get('guias', '')
        guias_list = [g.strip() for g in guias_input.strip().splitlines() if g.strip()]
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        mensajero_obj = next((m for m in mensajeros if m.nombre == mensajero_nombre), None)

        if not mensajero_obj:
            flash('Mensajero no encontrado', 'danger')
            return redirect(url_for('despachar_guias'))

        zona_obj = mensajero_obj.zona
        errores, exito = [], []

        with get_conn() as conn:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            for numero in guias_list:
                # guía existe
                cur.execute("SELECT 1 FROM guias WHERE numero_guia = %s;", (numero,))
                if not cur.fetchone():
                    errores.append(f'Guía {numero} no existe (FALTANTE)')
                    continue
                # ya recepcionada?
                cur.execute("SELECT * FROM recepciones WHERE numero_guia = %s;", (numero,))
                recepcion_existente = cur.fetchone()
                if recepcion_existente:
                    errores.append(f"Guía {numero} ya fue {recepcion_existente['tipo']}")
                    continue
                # ya despachada?
                cur.execute("SELECT * FROM despachos WHERE numero_guia = %s;", (numero,))
                despacho_existente = cur.fetchone()
                if despacho_existente:
                    errores.append(f'Guía {numero} ya fue despachada a {despacho_existente["mensajero"]}')
                    continue
                # insertar
                cur.execute("""
                    INSERT INTO despachos(numero_guia, mensajero, zona, fecha)
                    VALUES (%s, %s, %s, %s);
                """, (numero, mensajero_nombre, zona_obj.nombre if zona_obj else None, fecha))
                exito.append(f'Guía {numero} despachada a {mensajero_nombre}')

        if errores:
            flash("Errores:<br>" + "<br>".join(errores), 'danger')
        if exito:
            flash("Despachos exitosos:<br>" + "<br>".join(exito), 'success')

        cargar_datos_desde_db()
        return redirect(url_for('ver_despacho'))

    return render_template('despachar_guias.html',
                           mensajeros=[m.nombre for m in mensajeros],
                           zonas=[z.nombre for z in zonas])

# ---------- Ver despachos + export ----------

@app.route("/ver_despacho")
def ver_despacho():
    mensa = (request.args.get('mensajero') or '').strip()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    sql = """
        SELECT
            DATE(d.fecha) AS fecha,
            d.mensajero,
            d.zona,
            COUNT(*) AS total_guias
        FROM despachos d
        WHERE 1=1
    """
    params = []
    if mensa:
        sql += " AND d.mensajero = %s"
        params.append(mensa)
    if fi:
        sql += " AND DATE(d.fecha) >= %s"
        params.append(fi)
    if ff:
        sql += " AND DATE(d.fecha) <= %s"
        params.append(ff)
    sql += " GROUP BY DATE(d.fecha), d.mensajero, d.zona ORDER BY DATE(d.fecha) DESC"

    resumen = db_fetchall_dict(sql, params=params)

    return render_template(
        'ver_despacho.html',
        resumen=resumen,
        mensajeros=[m.nombre for m in mensajeros],
        mensajero_sel=mensa,
        fi=fi, ff=ff
    )


@app.get("/ver_despacho/export")
def export_despacho():
    mensa = (request.args.get('mensajero') or '').strip()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    sql = """
        SELECT
            DATE(d.fecha) AS fecha,
            d.mensajero,
            d.zona,
            COUNT(*) AS total_guias
        FROM despachos d
        WHERE 1=1
    """
    params = []
    if mensa:
        sql += " AND d.mensajero = %s"
        params.append(mensa)
    if fi:
        sql += " AND DATE(d.fecha) >= %s"
        params.append(fi)
    if ff:
        sql += " AND DATE(d.fecha) <= %s"
        params.append(ff)
    sql += " GROUP BY DATE(d.fecha), d.mensajero, d.zona ORDER BY DATE(d.fecha) DESC"

    df = read_sql_df(sql, params=params)
    return df_to_excel_download(df, base_name="despachos_resumen", sheet_name="Resumen")

# ---------- NUEVO: Pendiente (despachadas sin gestión) + export ----------

@app.route("/pendiente")
def pendiente():
    mensa = (request.args.get('mensajero') or '').strip()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    # Despachos sin recepción
    sql = """
        SELECT
            d.numero_guia,
            d.mensajero,
            d.zona,
            d.fecha     AS fecha_asignada,
            g.remitente,
            g.destinatario,
            g.direccion,
            g.ciudad
        FROM despachos d
        LEFT JOIN recepciones r ON r.numero_guia = d.numero_guia
        LEFT JOIN guias g       ON g.numero_guia = d.numero_guia
        WHERE r.numero_guia IS NULL
    """
    params = []
    if mensa:
        sql += " AND d.mensajero = %s"
        params.append(mensa)
    if fi:
        sql += " AND DATE(d.fecha) >= %s"
        params.append(fi)
    if ff:
        sql += " AND DATE(d.fecha) <= %s"
        params.append(ff)
    sql += " ORDER BY d.fecha DESC"

    rows = db_fetchall_dict(sql, params=params)
    return render_template("pendiente.html",
                           rows=rows,
                           mensajeros=[m.nombre for m in mensajeros],
                           mensajero_sel=mensa,
                           fi=fi, ff=ff)

@app.get("/pendiente/export")
def pendiente_export():
    mensa = (request.args.get('mensajero') or '').strip()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    sql = """
        SELECT
            d.numero_guia,
            d.mensajero,
            d.zona,
            d.fecha     AS fecha_asignada,
            g.remitente,
            g.destinatario,
            g.direccion,
            g.ciudad
        FROM despachos d
        LEFT JOIN recepciones r ON r.numero_guia = d.numero_guia
        LEFT JOIN guias g       ON g.numero_guia = d.numero_guia
        WHERE r.numero_guia IS NULL
    """
    params = []
    if mensa:
        sql += " AND d.mensajero = %s"
        params.append(mensa)
    if fi:
        sql += " AND DATE(d.fecha) >= %s"
        params.append(fi)
    if ff:
        sql += " AND DATE(d.fecha) <= %s"
        params.append(ff)
    sql += " ORDER BY d.fecha DESC"

    df = read_sql_df(sql, params=params)
    return df_to_excel_download(df, base_name="pendiente", sheet_name="Pendiente")

# ---------- Registrar / ver recepciones + export ----------

@app.route("/registrar_recepcion", methods=["GET", "POST"])
def registrar_recepcion():
    if request.method == 'POST':
        numero_guia = request.form.get('numero_guia')
        tipo = request.form.get('estado')  # ENTREGADA o DEVUELTA
        motivo = request.form.get('motivo', '')
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        existe_guia = db_fetchone_dict("SELECT 1 AS x FROM guias WHERE numero_guia = %s;", (numero_guia,))
        if not existe_guia:
            flash('Número de guía no existe en la base (FALTANTE)', 'danger')
            return redirect(url_for('registrar_recepcion'))

        despacho_existente = db_fetchone_dict("SELECT * FROM despachos WHERE numero_guia = %s;", (numero_guia,))
        if not despacho_existente:
            flash('La guía no ha sido despachada aún', 'warning')
            return redirect(url_for('registrar_recepcion'))

        recepcion_existente = db_fetchone_dict("SELECT 1 AS x FROM recepciones WHERE numero_guia = %s;", (numero_guia,))
        if recepcion_existente:
            flash('La recepción para esta guía ya está registrada', 'warning')
            return redirect(url_for('registrar_recepcion'))

        db_exec("""
            INSERT INTO recepciones(numero_guia, tipo, motivo, fecha)
            VALUES (%s, %s, %s, %s);
        """, (numero_guia, tipo, (motivo if tipo == 'DEVUELTA' else ''), fecha))

        flash(f'Recepción de guía {numero_guia} registrada como {tipo}', 'success')
        cargar_datos_desde_db()
        return redirect(url_for('registrar_recepcion'))

    return render_template('registrar_recepcion.html')

@app.route("/ver_recepciones")
def ver_recepciones():
    numero = (request.args.get('numero_guia') or '').strip().lower()
    tipo = (request.args.get('tipo') or '').strip().upper()  # ENTREGADA/DEVUELTA
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    lista = recepciones
    if numero:
        lista = [r for r in lista if numero in (r['numero_guia'] or '').lower()]
    if tipo:
        lista = [r for r in lista if (r['tipo'] or '').upper() == tipo]
    if fi:
        lista = [r for r in lista if str(r['fecha'])[:10] >= fi]
    if ff:
        lista = [r for r in lista if str(r['fecha'])[:10] <= ff]

    return render_template('ver_recepciones.html', recepciones=list(lista))

@app.get("/ver_recepciones/export")
def export_recepciones():
    numero = (request.args.get('numero_guia') or '').strip().lower()
    tipo = (request.args.get('tipo') or '').strip().upper()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()

    sql = """
        SELECT numero_guia, tipo, motivo, fecha
        FROM recepciones
        WHERE 1=1
    """
    params = []
    if numero:
        sql += " AND LOWER(COALESCE(numero_guia,'')) LIKE %s"
        params.append(f"%{numero}%")
    if tipo:
        sql += " AND UPPER(COALESCE(tipo,'')) = %s"
        params.append(tipo)
    if fi:
        sql += " AND DATE(fecha) >= %s"
        params.append(fi)
    if ff:
        sql += " AND DATE(fecha) <= %s"
        params.append(ff)
    sql += " ORDER BY fecha DESC"

    df = read_sql_df(sql, params=params)
    return df_to_excel_download(df, base_name="recepciones", sheet_name="Recepciones")

# ---------- Consulta estado ----------

@app.route("/consultar_estado", methods=["GET", "POST"])
def consultar_estado():
    resultado = None
    if request.method == 'POST':
        numero_guia = request.form.get('numero_guia', '').strip()
        if not numero_guia:
            flash('Debe ingresar un número de guía', 'warning')
            return redirect(url_for('consultar_estado'))

        existe_guia = db_fetchone_dict("SELECT 1 AS x FROM guias WHERE numero_guia = %s;", (numero_guia,))
        if not existe_guia:
            resultado = {'numero_guia': numero_guia, 'estado': 'FALTANTE', 'motivo': '', 'mensajero': '', 'zona': '', 'fecha': ''}
        else:
            despacho = db_fetchone_dict("SELECT * FROM despachos WHERE numero_guia = %s;", (numero_guia,))
            recepcion = db_fetchone_dict("SELECT * FROM recepciones WHERE numero_guia = %s;", (numero_guia,))

            if recepcion:
                estado = recepcion['tipo']
                motivo = recepcion['motivo']
                mensajero = despacho['mensajero'] if despacho else ''
                zona = despacho['zona'] if despacho else ''
                fecha = recepcion['fecha']
            elif despacho:
                estado = 'DESPACHADA'
                motivo = ''
                mensajero = despacho['mensajero']
                zona = despacho['zona']
                fecha = despacho['fecha']
            else:
                estado = 'EN VERIFICACION'
                motivo = ''
                mensajero = ''
                zona = ''
                fecha = ''

            resultado = {
                'numero_guia': numero_guia,
                'estado': estado,
                'motivo': motivo,
                'mensajero': mensajero,
                'zona': zona,
                'fecha': fecha
            }
    return render_template('consultar_estado.html', resultado=resultado)

# ---------- Liquidación + export ----------

@app.route("/liquidacion", methods=["GET", "POST"])
def liquidacion():
    liquidacion = None
    if request.method == 'POST':
        mensajero_nombre = request.form.get('mensajero')
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')

        try:
            datetime.strptime(fecha_inicio, '%Y-%m-%d')
            datetime.strptime(fecha_fin,  '%Y-%m-%d')
        except Exception:
            flash('Formato de fechas inválido', 'danger')
            return redirect(url_for('liquidacion'))

        gui_despachadas = db_fetchall_dict(
            'SELECT * FROM despachos WHERE mensajero = %s AND DATE(fecha) BETWEEN %s AND %s;',
            (mensajero_nombre, fecha_inicio, fecha_fin)
        )

        cantidad_guias = len(gui_despachadas)
        mensajero_obj = next((m for m in mensajeros if m.nombre == mensajero_nombre), None)
        tarifa = mensajero_obj.zona.tarifa if mensajero_obj and mensajero_obj.zona else 0
        total_pagar = cantidad_guias * tarifa

        liquidacion = {
            'mensajero': mensajero_nombre,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'cantidad_guias': cantidad_guias,
            'tarifa': tarifa,
            'total_pagar': total_pagar
        }
    return render_template('liquidacion.html', mensajeros=mensajeros, liquidacion=liquidacion)

@app.get("/liquidacion/export")
def export_liquidacion():
    mensajero_nombre = (request.args.get('mensajero') or '').strip()
    fecha_inicio = (request.args.get('fecha_inicio') or '').strip()
    fecha_fin = (request.args.get('fecha_fin') or '').strip()

    if not mensajero_nombre or not fecha_inicio or not fecha_fin:
        df_empty = pd.DataFrame(columns=["mensajero", "fecha_inicio", "fecha_fin", "cantidad_guias", "tarifa", "total_pagar"])
        return df_to_excel_download(df_empty, base_name="liquidacion", sheet_name="Resumen")

    df_detalle = read_sql_df("""
        SELECT numero_guia, mensajero, zona, fecha
        FROM despachos
        WHERE mensajero = %s AND DATE(fecha) BETWEEN %s AND %s
        ORDER BY fecha DESC
    """, params=[mensajero_nombre, fecha_inicio, fecha_fin])

    cantidad_guias = len(df_detalle)
    mensajero_obj = next((m for m in mensajeros if m.nombre == mensajero_nombre), None)
    tarifa = mensajero_obj.zona.tarifa if mensajero_obj and mensajero_obj.zona else 0
    total_pagar = cantidad_guias * tarifa

    df_resumen = pd.DataFrame([{
        "mensajero": mensajero_nombre,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "cantidad_guias": cantidad_guias,
        "tarifa": tarifa,
        "total_pagar": total_pagar
    }])

    # Excel con dos hojas
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Resumen
        df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
        ws = writer.sheets["Resumen"]
        for idx, col in enumerate(df_resumen.columns, start=1):
            max_len = max([len(str(col))] + [len(str(x)) for x in df_resumen[col].astype(str).values])
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(40, max_len + 2))
        # Detalle
        df_detalle2 = df_detalle.copy()
        for col in df_detalle2.columns:
            if pd.api.types.is_datetime64_any_dtype(df_detalle2[col]):
                try:
                    df_detalle2[col] = pd.to_datetime(df_detalle2[col]).dt.tz_localize(None)
                except Exception:
                    pass
        df_detalle2.to_excel(writer, index=False, sheet_name="Detalle")
        ws2 = writer.sheets["Detalle"]
        for idx, col in enumerate(df_detalle2.columns, start=1):
            max_len = max([len(str(col))] + [len(str(x)) for x in df_detalle2[col].astype(str).values]) if not df_detalle2.empty else len(str(col))
            ws2.column_dimensions[get_column_letter(idx)].width = max(12, min(40, max_len + 2))
        if "fecha" in df_detalle2.columns:
            col_idx = list(df_detalle2.columns).index("fecha") + 1
            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws2.max_row):
                for cell in row:
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"liquidacion_{stamp}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------- NUEVO: Clientes (crear/listar: SOLO NOMBRE) ----------

@app.route("/clientes", methods=["GET", "POST"])
def clientes_view():
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()

        if not nombre:
            flash("El nombre del cliente es obligatorio.", "danger")
            return redirect(url_for("clientes_view"))

        ya = db_fetchone_dict("SELECT 1 AS x FROM clientes WHERE LOWER(nombre)=LOWER(%s);", (nombre,))
        if ya:
            flash("Ese cliente ya existe.", "warning")
        else:
            # Solo nombre por ahora; los demás campos quedan NULL
            db_exec("INSERT INTO clientes(nombre) VALUES (%s);", (nombre,))
            flash("Cliente creado.", "success")

        cargar_datos_desde_db()
        return redirect(url_for("clientes_view"))

    return render_template("clientes.html", clientes=clientes)

# ---------- Recogidas + export (con cliente_id) ----------

@app.route("/registrar_recogida", methods=["GET", "POST"])
def registrar_recogida():
    if request.method == 'POST':
        numero_guia = request.form.get('numero_guia', '').strip()
        fecha = request.form.get('fecha')
        observaciones = request.form.get('observaciones', '').strip()
        cliente_id = request.form.get('cliente_id')  # puede venir vacío

        if not numero_guia or not fecha:
            flash('Debe completar número de guía y fecha', 'danger')
            return redirect(url_for('registrar_recogida'))

        # Normaliza cliente_id
        cliente_id_val = int(cliente_id) if (cliente_id and cliente_id.isdigit()) else None

        db_exec("""
            INSERT INTO recogidas(numero_guia, fecha, observaciones, cliente_id)
            VALUES (%s, %s, %s, %s);
        """, (numero_guia, fecha, observaciones, cliente_id_val))

        flash(f'Recogida registrada para la guía {numero_guia}', 'success')
        cargar_datos_desde_db()
        return redirect(url_for('registrar_recogida'))

    return render_template('registrar_recogida.html', clientes=clientes)

# ---- (Opcional) Alta rápida de cliente desde “Registrar recogida” ----

@app.post("/clientes_quick")
def clientes_quick():
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        flash("El nombre del cliente es obligatorio.", "danger")
        return redirect(url_for("registrar_recogida"))

    ya = db_fetchone_dict("SELECT 1 AS x FROM clientes WHERE LOWER(nombre)=LOWER(%s);", (nombre,))
    if ya:
        flash("Ese cliente ya existe.", "warning")
    else:
        db_exec("INSERT INTO clientes(nombre) VALUES (%s);", (nombre,))
        flash("Cliente creado.", "success")

    cargar_datos_desde_db()
    return redirect(url_for("registrar_recogida"))

@app.route("/ver_recogidas")
def ver_recogidas():
    filtro_numero = (request.args.get('filtro_numero') or '').strip().lower()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()
    cliente_id = (request.args.get('cliente_id') or '').strip()

    # Traemos ya unido con clientes y aplicamos filtros en SQL
    sql = """
        SELECT
            r.id,
            r.numero_guia,
            r.fecha,
            r.observaciones,
            r.cliente_id,
            c.nombre AS cliente
        FROM recogidas r
        LEFT JOIN clientes c ON c.id = r.cliente_id
        WHERE 1=1
    """
    params = []

    if filtro_numero:
        sql += " AND LOWER(COALESCE(r.numero_guia,'')) LIKE %s"
        params.append(f"%{filtro_numero}%")

    if fi:
        sql += " AND DATE(r.fecha) >= %s"
        params.append(fi)

    if ff:
        sql += " AND DATE(r.fecha) <= %s"
        params.append(ff)

    if cliente_id and cliente_id.isdigit():
        sql += " AND r.cliente_id = %s"
        params.append(int(cliente_id))

    sql += " ORDER BY r.fecha DESC"

    rows = db_fetchall_dict(sql, params=params)

    return render_template(
        'ver_recogidas.html',
        recogidas=rows,
        clientes=clientes,          # para el select
        cliente_sel=cliente_id,     # para mantener selección
        fi=fi, ff=ff,
        filtro_numero=(request.args.get('filtro_numero') or '').strip()
    )


@app.get("/ver_recogidas/export")
def export_recogidas():
    filtro_numero = (request.args.get('filtro_numero') or '').strip().lower()
    fi = (request.args.get('fi') or '').strip()
    ff = (request.args.get('ff') or '').strip()
    cliente_id = (request.args.get('cliente_id') or '').strip()

    sql = """
        SELECT
            r.id,
            r.numero_guia,
            r.fecha,
            r.observaciones,
            c.nombre AS cliente
        FROM recogidas r
        LEFT JOIN clientes c ON c.id = r.cliente_id
        WHERE 1=1
    """
    params = []

    if filtro_numero:
        sql += " AND LOWER(COALESCE(r.numero_guia, '')) LIKE %s"
        params.append(f"%{filtro_numero}%")

    if fi:
        sql += " AND DATE(r.fecha) >= %s"
        params.append(fi)

    if ff:
        sql += " AND DATE(r.fecha) <= %s"
        params.append(ff)

    if cliente_id and cliente_id.isdigit():
        sql += " AND r.cliente_id = %s"
        params.append(int(cliente_id))

    sql += " ORDER BY r.fecha DESC"

    df = read_sql_df(sql, params=params)
    if df.empty:
        df = pd.DataFrame(columns=["id", "numero_guia", "fecha", "observaciones", "cliente"])

    return df_to_excel_download(df, base_name="recogidas", sheet_name="Recogidas")

# ---------- Endpoints util ----------

@app.route("/health")
def health():
    db_fetchone_dict("SELECT 1 AS ok;")
    return jsonify(ok=True)

@app.route("/init")
def init():
    ensure_schema()
    new_id = db_fetchone_dict(
        "INSERT INTO guias(remitente, numero_guia, destinatario, direccion, ciudad) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (numero_guia) DO NOTHING RETURNING numero_guia;",
        ("demo", "DUMMY-0001", "destino", "direccion", "ciudad")
    )
    return jsonify(ok=True, demo_insert=new_id)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
